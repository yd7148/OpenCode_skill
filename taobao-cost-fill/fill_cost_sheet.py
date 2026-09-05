#!/usr/bin/env python3
"""fill_cost_sheet.py — 依訂單卡片清單填入「淘寶費用計算明細」R0 樣板並存成 R1。

用法:
  .venv/bin/python fill_cost_sheet.py \
      --in   <R0.xlsx> \
      --out  <R1.xlsx>        (省略時自動 *_R0.xlsx -> *_R1.xlsx) \
      --sheet <分頁名>         (省略時自動抓檔名中的日期) \
      --cards <cards.md>      (省略時由 stdin 讀取)

卡片行格式:
  ## #02商品名稱  實付金額 ￥16.00  极兔速递 JT3175106025237  JT3175106025237 台灣 0.920
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl

START_ROW = 4  # row4 起為第一筆商品（項次=1）
DEFAULT_SHEET_RE = re.compile(r"(19|20)\d{2}-\d{2}-\d{2}")
CARD_RE = re.compile(r"^##\s*#(\d+)\s*(.*?)\s*實付金額\s*￥([\d.]+)(.*)$", re.S)


def parse_card(line: str):
    m = CARD_RE.match(line.strip())
    if not m:
        print(f"WARN 無法解析卡片行，略過: {line!r}", file=sys.stderr)
        return None
    num = int(m.group(1))
    name = m.group(2).strip()
    price = float(m.group(3))
    rest = m.group(4) or ""
    wm = re.search(r"台灣\s*([\d.]+)\s*$", rest)
    weight = float(wm.group(1)) if wm else None
    return num, name, price, weight


def pick_sheet(wb, sheet_arg, in_path):
    if sheet_arg:
        if sheet_arg not in wb.sheetnames:
            sys.exit(f"找不到分頁: {sheet_arg}（現有: {wb.sheetnames}）")
        return sheet_arg
    dm = DEFAULT_SHEET_RE.search(Path(in_path).name)
    if dm:
        for s in wb.sheetnames:
            if s.strip() == dm.group(0).strip():
                return s
    sys.exit(f"無法決定分頁，請用 --sheet 指定。現有: {wb.sheetnames}")


def structural_rows(ws, n_max):
    """找出運費/關稅/合計等結構列（回傳其列號，依序）。"""
    out = []
    for r in range(START_ROW, n_max + 1):
        c = ws.cell(row=r, column=3).value
        g = ws.cell(row=r, column=7).value
        if (g == "運費") or (isinstance(c, str) and ("達飛" in c or "达飞" in c or c.strip() == "賴政府關稅" or c.strip() == "合計")):
            out.append(r)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--out", dest="dst", default=None)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--cards", default=None)
    args = ap.parse_args()

    if args.dst is None:
        args.dst = re.sub(r"-R0\.xlsx$", "-R1.xlsx", args.src)
        if args.dst == args.src:
            sys.exit("無法自動產生 R1 檔名（來源檔名不含 -R0.xlsx），請用 --out 指定。")

    cards = []
    if args.cards:
        lines = Path(args.cards).read_text(encoding="utf-8").splitlines()
    else:
        lines = sys.stdin.read().splitlines()
    for line in lines:
        if not line.strip():
            continue
        c = parse_card(line)
        if c:
            cards.append(c)
    if not cards:
        sys.exit("沒有解析到任何訂單卡片。")
    cards.sort()
    print(f"解析到 {len(cards)} 筆訂單（# {cards[0][0]}..#{cards[-1][0]}）", file=sys.stderr)

    wb = openpyxl.load_workbook(args.src)
    ws = wb[pick_sheet(wb, args.sheet, args.src)]

    for idx, (num, name, price, weight) in enumerate(cards):
        r = START_ROW + idx
        ws.cell(row=r, column=2, value=num)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=price)
        ws.cell(row=r, column=8, value=weight if weight is not None else None)

    last_used = START_ROW + len(cards) - 1
    structs = structural_rows(ws, last_used + 40)
    boundary = min([r for r in structs if r > last_used], default=None)
    if boundary:
        for r in range(last_used + 1, boundary):
            if ws.cell(row=r, column=2).value is not None:
                ws.cell(row=r, column=2).value = None
        print(f"清除 {last_used + 1}..{boundary - 1} 的幽靈項次", file=sys.stderr)

    wb.save(args.dst)
    print(f"saved -> {args.dst}", file=sys.stderr)


if __name__ == "__main__":
    main()